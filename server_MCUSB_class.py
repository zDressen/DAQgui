from fractions import Fraction
from sys import stdout
from collections import namedtuple
from uldaq import (get_daq_device_inventory, DaqDevice, AInScanFlag,
                   DaqEventType, WaitType, ScanOption, InterfaceType,
                   AiInputMode, create_float_buffer, ULException,
                   EventCallbackArgs, AoInfo, AiInfo)
from uldaq import (get_daq_device_inventory, DaqDevice, create_float_buffer, InterfaceType, DaqOutScanFlag, Range,
                   ScanOption, DigitalDirection, DigitalPortType, ScanStatus, DaqOutChanType, DaqOutChanDescriptor)
import os
import h5py
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import argparse
from math import pi, sin, pow
from scipy.signal import square, sawtooth
import time
import zmq
from time import sleep
import string

class MCUSBClass:
    #parser = argparse.ArgumentParser(description='reads MCCDAQ into .hdf file')
    #parser.add_argument('-f','--file', type=str, required=True, help='Filename')
    #args = parser.parse_args()
    #OUTPUT_FILENAME = args.file
    OUTPUT_FILENAME = 'test.hdf'
    f_out = None
    dset = None
    # File output
    #columnNames = ['chan0','chan1','chan2', 'chan3','chan4','chan5','chan6','chan7']
    columnNames = ['chan0','chan1']
    #hdfKey = 'events'
    #comp_lib = 'bzip2'   # compression library
    #comp_level = 9       # compression level [0-9], 0 disables compression
    
    PLOT_ON = True
    low_channel = 0
    high_channel = len(columnNames) - 1
    samples_per_channel = 3000 # Samples per channel in buffer
    rate = 1000              # Daq rate [Hz]
    buffer_store = 2000  # Store buffer into HDF file after this many samples
                         # Cannot be > samples_per_channel
                         
    # Continuous vs finite scan
    #scan_options = ScanOption.DEFAULTIO
    scan_options = ScanOption.CONTINUOUS

    # Single ended vs differential input
    input_mode = AiInputMode.SINGLE_ENDED #Box uses AC coupling or sumn
    #input_mode= AiInputMode.DIFFERENTIAL
    
    event_types = DaqEventType.ON_DATA_AVAILABLE | DaqEventType.ON_END_OF_INPUT_SCAN | DaqEventType.ON_INPUT_SCAN_ERROR
    
    flags = AInScanFlag.DEFAULT
    
    def __init__(mysillyobject, daq_device, ai_device, ao_device, data, out_buffer, ranges, range_index):
        mysillyobject.daq_device = daq_device
        mysillyobject.ai_device = ai_device
        mysillyobject.data = data
        mysillyobject.out_buffer = out_buffer
        mysillyobject.ranges = ranges
        mysillyobject.range_index = range_index
        mysillyobject.ao_device = ao_device

    
DAQ = MCUSBClass(None, None, None, 0, 0, 0, 0)

def MCUSB_initiate():    
    ai_device = DAQ.ai_device

    PLOT_ON = DAQ.PLOT_ON
    low_channel = DAQ.low_channel
    high_channel = DAQ.high_channel
    samples_per_channel = DAQ.samples_per_channel
    rate = DAQ.rate
    buffer_store = DAQ.buffer_store
    flags = DAQ.flags
    input_mode = DAQ.input_mode
    scan_options = DAQ.scan_options
    event_types = DAQ.event_types

    descriptor_index = 0
    range_index = 0
    interface_type = InterfaceType.USB
    ScanParams = namedtuple('ScanParams', 'buffer high_chan low_chan buffer_store')
    successFlag = True # for more graceful error handling of ctrl C
    
    os.system('clear')
    
    #--------------------------------------------------    
    # Get descriptors for all of the available DAQ devices.
    devices = get_daq_device_inventory(interface_type)
    number_of_devices = len(devices)
    if number_of_devices == 0:
        raise Exception('Error: No DAQ devices found')

    print('Found', number_of_devices, 'DAQ device(s):')
    for i in range(number_of_devices):
        print('  ', devices[i].product_name, ' (', devices[i].unique_id, ')', sep='')

    # Create the DAQ device object associated with the specified descriptor index.
    DAQ.daq_device = DaqDevice(devices[descriptor_index])

    # Get the AiDevice object and verify that it is valid.
    DAQ.ai_device = DAQ.daq_device.get_ai_device()
    if DAQ.ai_device is None:
        raise Exception('Error: The DAQ device does not support analog input')
        
    # Get the AoDevice object and verify that it is valid.
    DAQ.ao_device = DAQ.daq_device.get_ao_device()
    if DAQ.ao_device is None:
        raise Exception('Error: The DAQ device does not support analog output')

    # Verify that the specified device supports hardware pacing for analog input.
    ai_info = DAQ.ai_device.get_info()
    print('scan options = ', ai_info.get_scan_options())
    
    if not ai_info.has_pacer():
        raise Exception('\nError: The specified DAQ device does not support hardware paced analog input')
    
    #MCUSB_ao()
    #-------------------------------------------------------
     
    # Establish a connection to the DAQ device.
    descriptor = DAQ.daq_device.get_descriptor()
    print('\nConnecting to', descriptor.dev_string, '- ...')
    DAQ.daq_device.connect()

    # Get the number of channels and validate the high channel number.
    number_of_channels = ai_info.get_num_chans_by_mode(input_mode)
    if high_channel >= number_of_channels:
        high_channel = number_of_channels - 1
    channel_count = high_channel - low_channel + 1

    # Get a list of supported ranges and validate the range index.
    DAQ.ranges = ai_info.get_ranges(input_mode)
    if DAQ.range_index >= len(DAQ.ranges):
        DAQ.range_index = len(DAQ.ranges) - 1

    # Allocate a buffer to receive the data.
    DAQ.data = create_float_buffer(channel_count, samples_per_channel)

    # Store the scan event parameters for use in the callback function.
    scan_event_parameters = ScanParams(DAQ.data, high_channel, low_channel, buffer_store)

    # Enable the event to be notified every time samples are available.
    DAQ.daq_device.enable_event(event_types, buffer_store, event_callback_function,
                                scan_event_parameters)

    print('\n', descriptor.dev_string, ' is ready', sep='')
    print('    Channels: ', low_channel, '-', high_channel)
    print('    Input mode: ', input_mode.name)
    print('    Range: ', DAQ.ranges[DAQ.range_index].name)
    print('    Samples per channel: ', samples_per_channel)
    print('    Rate: ', rate, 'Hz')
    print('    Scan options:', display_scan_options(scan_options))
    print('Waiting for requests from the client...')
    
    return
    
def MCUSB_ao(waveform_stats):
    print("Forming wave!")
    #Take waveform_stats useful info for generating data
    wave_info = []
    file_name = None
    sample_rate = 0
    samples_per_channel = 0
    samples_per_cycle = 0
    analog_low_channel = 0
    analog_high_channel = 0
    scan_options = ScanOption.CONTINUOUS
    if waveform_stats[0] == 'Custom':
        samples_per_channel = waveform_stats[2]
        sample_rate = waveform_stats[3]
        file_name = waveform_stats[1]
        samples_per_cycle = waveform_stats[2]
        analog_low_channel = waveform_stats[5]
        analog_high_channel = waveform_stats[4]
        if waveform_stats[5] == "Single":
            scan_options = ScanOption.SINGLEIO
    else:
        wave_info = [waveform_stats[0],int(waveform_stats[1]),float(waveform_stats[2]),float(waveform_stats[3]),float(waveform_stats[4])]
        sample_rate = int(waveform_stats[5])  # Hz
        samples_per_channel = sample_rate * int(waveform_stats[1])
        samples_per_cycle = int(sample_rate * int(wave_info[1]))
        analog_low_channel = int(waveform_stats[7])
        analog_high_channel = int(waveform_stats[6])
        if waveform_stats[8] == "Single":
            scan_options = ScanOption.SINGLEIO
    channel_descriptors = []
    scan_flags = DaqOutScanFlag.DEFAULT

    # Parameters used when creating channel_descriptors list
    analog_range_index = 0
    digital_low_port_index = 0
    digital_high_port_index = 0
    
    #-----------------------------------------------------
    # Create the daq output object and verify that it is valid
    daqo_device = DAQ.daq_device.get_daqo_device()
    # Verify the specified DAQ device supports DAQ output.
    if daqo_device is None:
        raise Exception('Error: The DAQ device does not support DAQ output')

    daqo_info = daqo_device.get_info()
    
    
    descriptor = DAQ.daq_device.get_descriptor()
    #print('\nConnecting to', descriptor.dev_string, '- please wait...')
    
    # Configure supported analog input and digital input channels
    amplitudes = []
    supported_channel_types = daqo_info.get_channel_types()
    if DaqOutChanType.ANALOG in supported_channel_types:
        configure_analog_channels(DAQ.daq_device, analog_low_channel, analog_high_channel,
                                    analog_range_index, channel_descriptors, amplitudes)
    if DaqOutChanType.DIGITAL in supported_channel_types:
        configure_digital_channels(DAQ.daq_device, digital_low_port_index, digital_high_port_index,
                                    channel_descriptors, amplitudes)

    num_channels = len(channel_descriptors)

    # Create a buffer for output data.
    DAQ.out_buffer = create_float_buffer(num_channels, samples_per_channel)
    # Fill the output buffer with data.
    if waveform_stats[0] == 'Custom':
        create_output_data_xsl(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info)
    else:
        create_output_data(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info)

    print('\n', descriptor.dev_string, 'ready')
    print('    Function demonstrated: DaqoDevice.daq_out_scan')
    print('    Number of Scan Channels:', num_channels)
    for chan in range(num_channels):
        chan_descriptor = channel_descriptors[chan]  # type: DaqOutChanDescriptor
        print('        Scan Channel', chan, end='')
        print(': type =', DaqOutChanType(chan_descriptor.type).name, end='')
        if chan_descriptor.type == DaqOutChanType.ANALOG:
            print(', channel =', chan_descriptor.channel, end='')
            print(', range =', Range(chan_descriptor.range).name, end='')
        else:
            print(', port =', DigitalPortType(chan_descriptor.channel).name, end='')
        print('')
    print('    Samples per channel:', samples_per_channel)
    print('    Rate:', sample_rate, 'Hz')
    print('    Scan options:', display_scan_options(scan_options))
    #try:
    #    input('\nHit ENTER to generate sin waves....')
    #except (NameError, SyntaxError):
    #    pass

    #system('clear')
    
    # Start the output scan.
    sample_rate = daqo_device.daq_out_scan(channel_descriptors, samples_per_channel, sample_rate,
                                            scan_options, scan_flags, DAQ.out_buffer)
    
    #print('\n*Press Ctrl-C to stop scan')
    print('\n  AO on:  Actual scan rate:   ', sample_rate, 'Hz')
    
    return
    
# This and digital should work, Needs tested on the RASPI
def MCUSB_multi_ao(waveform1,waveform2,waveform_info):
    wave_info = []
    wave_info2 = []
    sample_rate = 0
    samples_per_channel = 0
    samples_per_cycle = 0
    analog_low_channel = 0
    analog_high_channel = 2
    scan_options = ScanOption.CONTINUOUS
    # Here we will separate waves into their own lists, making the process for filling the buffer
    # much easier. Additionally, the frequencies will be matched such that both waves are output
    # continuously without any major jumps. This process has been minorly adapted from Calvin Nettelhorst's
    # gui code for outputting dual waves.
    if waveform1 == "Custom":
        if waveform2 == "Custom":
            # Both waveform are custom
            wave_info.append("Custom").append(waveform_info[1])
            wave_info2.append("Custom").append(waveform_info[4])
            # Set Frequency
            frequency_ch1 = float(waveform_info[2] * np.pi * 2)
            frequency_ch2 = float(waveform_info[5] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
        else:
            # Wave 1 is custom, Wave 2 is a Preset
            wave_info.append("Custom").append(waveform_info[1])
            wave_info2 = wave_info[3:7]
            # Set Frequency
            frequency_ch1 = float(waveform_info[2] * np.pi * 2)
            frequency_ch2 = float(waveform_info[4] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
    else:
        if waveform2 == "Custom":
            # Wave 1 is  a Preset, Wave 2 is Custom
            wave_info = waveform_info[0:4]
            wave_info2.append("Custom").append(waveform_info[6])
            # Set Frequency
            frequency_ch1 = float(waveform_info[1] * np.pi * 2)
            frequency_ch2 = float(waveform_info[7] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
        else:
            # Wave 1 and 2 are both Presets
            sample_rate = waveform_info[10]
            wave_info = waveform_info[0:4]
            wave_info2 = waveform_info[5:9]
            frequency_ch1 = float(waveform_info[1] * np.pi * 2)
            frequency_ch2 = float(waveform_info[6] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
            if wave_info[13] == "Single":
                scan_options = ScanOption.SINGLEIO
            samples_per_cycle = samples_per_channel * 2
    channel_descriptors = []
    scan_flags = DaqOutScanFlag.DEFAULT

    # Parameters used when creating channel_descriptors list
    analog_range_index = 0
    digital_low_port_index = 0
    digital_high_port_index = 0
    
    #-----------------------------------------------------
    # Create the daq output object and verify that it is valid
    daqo_device = DAQ.daq_device.get_daqo_device()
    # Verify the specified DAQ device supports DAQ output.
    if daqo_device is None:
        raise Exception('Error: The DAQ device does not support DAQ output')

    daqo_info = daqo_device.get_info()
    
    
    descriptor = DAQ.daq_device.get_descriptor()
    #print('\nConnecting to', descriptor.dev_string, '- please wait...')
    
    # Configure supported analog input and digital input channels
    amplitudes = []
    supported_channel_types = daqo_info.get_channel_types()
    if DaqOutChanType.ANALOG in supported_channel_types:
        configure_analog_channels(DAQ.daq_device, analog_low_channel, analog_high_channel,
                                    analog_range_index, channel_descriptors, amplitudes)
    if DaqOutChanType.DIGITAL in supported_channel_types:
        configure_digital_channels(DAQ.daq_device, digital_low_port_index, digital_high_port_index,
                                    channel_descriptors, amplitudes)

    num_channels = len(channel_descriptors)
    # Create a buffer for output data.
    DAQ.out_buffer = create_float_buffer(num_channels, samples_per_channel)

    # Fill the buffer with data
    create_output_data_multi(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info, wave_info2)

    print('\n', descriptor.dev_string, 'ready')
    print('    Function demonstrated: DaqoDevice.daq_out_scan')
    print('    Number of Scan Channels:', num_channels)
    for chan in range(num_channels):
        chan_descriptor = channel_descriptors[chan]  # type: DaqOutChanDescriptor
        print('        Scan Channel', chan, end='')
        print(': type =', DaqOutChanType(chan_descriptor.type).name, end='')
        if chan_descriptor.type == DaqOutChanType.ANALOG:
            print(', channel =', chan_descriptor.channel, end='')
            print(', range =', Range(chan_descriptor.range).name, end='')
        else:
            print(', port =', DigitalPortType(chan_descriptor.channel).name, end='')
        print('')
    print('    Samples per channel:', samples_per_channel)
    print('    Rate:', sample_rate, 'Hz')
    print('    Scan options:', display_scan_options(scan_options))
    
    sample_rate = daqo_device.daq_out_scan(channel_descriptors, samples_per_channel, sample_rate,
                                            scan_options, scan_flags, DAQ.out_buffer)
    
    print('\n  AO on:  Actual scan rate:   ', sample_rate, 'Hz')
    
    return


# The Digital Output functions are exactly that of the ao functions, except the role of the channels has been swapped
# i.e. the analog_low_channel is not digital_low_port_index, functionally changing from analog to digital output
def MCUSB_do(waveform_stats):
    print("Forming wave!")
    #Take waveform_stats useful info for generating data
    wave_info = []
    file_name = None
    sample_rate = 0
    samples_per_channel = 0
    samples_per_cycle = 0
    digital_low_port_index = 0
    digital_high_port_index = 0
    scan_options = ScanOption.CONTINUOUS
    if waveform_stats[0] == 'Custom':
        samples_per_channel = waveform_stats[2]
        sample_rate = waveform_stats[3]
        file_name = waveform_stats[1]
        samples_per_cycle = waveform_stats[2]
        digital_low_port_index = waveform_stats[5]
        digital_high_port_index = waveform_stats[4]
        if waveform_stats[5] == "Single":
            scan_options = ScanOption.SINGLEIO
    else:
        wave_info = [waveform_stats[0],int(waveform_stats[1]),float(waveform_stats[2]),float(waveform_stats[3]),float(waveform_stats[4])]
        sample_rate = int(waveform_stats[5])  # Hz
        samples_per_channel = sample_rate * int(waveform_stats[1])
        samples_per_cycle = int(sample_rate * int(wave_info[1]))
        digital_low_port_index = int(waveform_stats[7])
        digital_high_port_index = int(waveform_stats[6])
        if waveform_stats[8] == "Single":
            scan_options = ScanOption.SINGLEIO
    channel_descriptors = []
    scan_flags = DaqOutScanFlag.DEFAULT

    # Parameters used when creating channel_descriptors list
    analog_range_index = 0
    analog_low_channel = 0
    analog_high_channel = 0
    
    #-----------------------------------------------------
    # Create the daq output object and verify that it is valid
    daqo_device = DAQ.daq_device.get_daqo_device()
    # Verify the specified DAQ device supports DAQ output.
    if daqo_device is None:
        raise Exception('Error: The DAQ device does not support DAQ output')

    daqo_info = daqo_device.get_info()
    
    
    descriptor = DAQ.daq_device.get_descriptor()
    #print('\nConnecting to', descriptor.dev_string, '- please wait...')
    
    # Configure supported analog input and digital input channels
    amplitudes = []
    supported_channel_types = daqo_info.get_channel_types()
    if DaqOutChanType.ANALOG in supported_channel_types:
        configure_analog_channels(DAQ.daq_device, analog_low_channel, analog_high_channel,
                                    analog_range_index, channel_descriptors, amplitudes)
    if DaqOutChanType.DIGITAL in supported_channel_types:
        configure_digital_channels(DAQ.daq_device, digital_low_port_index, digital_high_port_index,
                                    channel_descriptors, amplitudes)

    num_channels = len(channel_descriptors)

    # Create a buffer for output data.
    DAQ.out_buffer = create_float_buffer(num_channels, samples_per_channel)
    # Fill the output buffer with data.
    if waveform_stats[0] == 'Custom':
        create_output_data_xsl(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info)
    else:
        create_output_data(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info)

    print('\n', descriptor.dev_string, 'ready')
    print('    Function demonstrated: DaqoDevice.daq_out_scan')
    print('    Number of Scan Channels:', num_channels)
    for chan in range(num_channels):
        chan_descriptor = channel_descriptors[chan]  # type: DaqOutChanDescriptor
        print('        Scan Channel', chan, end='')
        print(': type =', DaqOutChanType(chan_descriptor.type).name, end='')
        if chan_descriptor.type == DaqOutChanType.ANALOG:
            print(', channel =', chan_descriptor.channel, end='')
            print(', range =', Range(chan_descriptor.range).name, end='')
        else:
            print(', port =', DigitalPortType(chan_descriptor.channel).name, end='')
        print('')
    print('    Samples per channel:', samples_per_channel)
    print('    Rate:', sample_rate, 'Hz')
    print('    Scan options:', display_scan_options(scan_options))
    #try:
    #    input('\nHit ENTER to generate sin waves....')
    #except (NameError, SyntaxError):
    #    pass

    #system('clear')
    
    # Start the output scan.
    sample_rate = daqo_device.daq_out_scan(channel_descriptors, samples_per_channel, sample_rate,
                                            scan_options, scan_flags, DAQ.out_buffer)
    
    #print('\n*Press Ctrl-C to stop scan')
    print('\n  AO on:  Actual scan rate:   ', sample_rate, 'Hz')
    
    return
    
def MCUSB_multi_do(waveform1,waveform2,waveform_info):
    wave_info = []
    wave_info2 = []
    sample_rate = 0
    samples_per_channel = 0
    samples_per_cycle = 0
    digital_low_port_index = 0
    digital_high_port_index = 2 # Need to grab from the actual functions
    scan_options = ScanOption.CONTINUOUS
    if waveform1 == "Custom":
        if waveform2 == "Custom":
            # Both waveform are custom
            wave_info.append("Custom").append(waveform_info[1])
            wave_info2.append("Custom").append(waveform_info[4])
            # Set Frequency
            frequency_ch1 = float(waveform_info[2] * np.pi * 2)
            frequency_ch2 = float(waveform_info[5] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
        else:
            # Wave 1 is custom, Wave 2 is a Preset
            wave_info.append("Custom").append(waveform_info[1])
            wave_info2 = wave_info[3:7]
            # Set Frequency
            frequency_ch1 = float(waveform_info[2] * np.pi * 2)
            frequency_ch2 = float(waveform_info[4] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
    else:
        if waveform2 == "Custom":
            # Wave 1 is  a Preset, Wave 2 is Custom
            wave_info = waveform_info[0:4]
            wave_info2.append("Custom").append(waveform_info[6])
            # Set Frequency
            frequency_ch1 = float(waveform_info[1] * np.pi * 2)
            frequency_ch2 = float(waveform_info[7] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
        else:
            # Wave 1 and 2 are both Presets
            sample_rate = waveform_info[10]
            wave_info = waveform_info[0:4]
            wave_info2 = waveform_info[5:9]
            frequency_ch1 = float(waveform_info[1] * np.pi * 2)
            frequency_ch2 = float(waveform_info[6] * np.pi * 2)
            freq1_fraction = Fraction(frequency_ch1).limit_denominator()
            freq2_fraction = Fraction(frequency_ch2).limit_denominator()
            # Find the least common multiple of the denominators
            lcm_denominator = np.lcm(freq1_fraction.denominator, freq2_fraction.denominator)
            # Scale the fundamental period to achieve integer points for each wave
            if frequency_ch1 >= frequency_ch2:
                points_per_period_wave1 = int(sample_rate/frequency_ch1)
                points_per_period_wave2 = int((freq1_fraction.denominator / freq2_fraction.denominator) * points_per_period_wave1)
                list_length = np.lcm(points_per_period_wave1, points_per_period_wave2) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch1)
            else:
                points_per_period_wave2 = int(sample_rate/frequency_ch2)
                points_per_period_wave1 = int((freq2_fraction.denominator / freq1_fraction.denominator) * points_per_period_wave2)
                list_length = np.lcm(points_per_period_wave2, points_per_period_wave1) * lcm_denominator
                samples_per_channel = int(list_length*frequency_ch2)
            if wave_info[13] == "Single":
                scan_options = ScanOption.SINGLEIO
            samples_per_cycle = samples_per_channel * 2
    channel_descriptors = []
    scan_flags = DaqOutScanFlag.DEFAULT

    # Parameters used when creating channel_descriptors list
    analog_range_index = 0
    analog_low_channel = 0
    analog_high_channel = 0
    
    #-----------------------------------------------------
    # Create the daq output object and verify that it is valid
    daqo_device = DAQ.daq_device.get_daqo_device()
    # Verify the specified DAQ device supports DAQ output.
    if daqo_device is None:
        raise Exception('Error: The DAQ device does not support DAQ output')

    daqo_info = daqo_device.get_info()
    
    
    descriptor = DAQ.daq_device.get_descriptor()
    #print('\nConnecting to', descriptor.dev_string, '- please wait...')
    
    # Configure supported analog input and digital input channels
    amplitudes = []
    supported_channel_types = daqo_info.get_channel_types()
    if DaqOutChanType.ANALOG in supported_channel_types:
        configure_analog_channels(DAQ.daq_device, analog_low_channel, analog_high_channel,
                                    analog_range_index, channel_descriptors, amplitudes)
    if DaqOutChanType.DIGITAL in supported_channel_types:
        configure_digital_channels(DAQ.daq_device, digital_low_port_index, digital_high_port_index,
                                    channel_descriptors, amplitudes)

    num_channels = len(channel_descriptors)
    # Create a buffer for output data.
    DAQ.out_buffer = create_float_buffer(num_channels, samples_per_channel)

    # Fill the buffer with data
    create_output_data_multi(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, DAQ.out_buffer, wave_info, wave_info2)

    print('\n', descriptor.dev_string, 'ready')
    print('    Function demonstrated: DaqoDevice.daq_out_scan')
    print('    Number of Scan Channels:', num_channels)
    for chan in range(num_channels):
        chan_descriptor = channel_descriptors[chan]  # type: DaqOutChanDescriptor
        print('        Scan Channel', chan, end='')
        print(': type =', DaqOutChanType(chan_descriptor.type).name, end='')
        if chan_descriptor.type == DaqOutChanType.ANALOG:
            print(', channel =', chan_descriptor.channel, end='')
            print(', range =', Range(chan_descriptor.range).name, end='')
        else:
            print(', port =', DigitalPortType(chan_descriptor.channel).name, end='')
        print('')
    print('    Samples per channel:', samples_per_channel)
    print('    Rate:', sample_rate, 'Hz')
    print('    Scan options:', display_scan_options(scan_options))
    
    sample_rate = daqo_device.daq_out_scan(channel_descriptors, samples_per_channel, sample_rate,
                                            scan_options, scan_flags, DAQ.out_buffer)
    
    print('\n  AO on:  Actual scan rate:   ', sample_rate, 'Hz')
    
    return

def MCUSB_stop_ao():
    daqo_device = DAQ.daq_device.get_daqo_device()
    daqo_device.scan_stop()

def configure_analog_channels(daq_device, low_channel, high_channel, range_index, channel_descriptors, amplitudes):
    """
    Add analog output channels to the channel_descriptors list.

    Raises:
        Exception if a channel is not in range.
    """
    ao_device = daq_device.get_ao_device()
    ao_info = ao_device.get_info()

    # Validate the low_channel and high_channel values
    num_channels = ao_info.get_num_chans()
    valid_channels_string = 'valid channels are 0 - {0:d}'.format(num_channels - 1)
    if low_channel < 0 or low_channel >= num_channels:
        error_message = ' '.join(['Error: Invalid analog_low_channel selection,', valid_channels_string])
        raise Exception(error_message)
    if high_channel < 0 or high_channel >= num_channels:
        error_message = ' '.join(['Error: Invalid analog_high_channel selection,', valid_channels_string])
        raise Exception(error_message)

    # Validate the range_index value
    voltage_ranges = ao_info.get_ranges()
    if range_index < 0:
        range_index = 0
    elif range_index >= len(voltage_ranges):
        range_index = len(voltage_ranges) - 1

    voltage_range = voltage_ranges[range_index]

    # Create a channel descriptor for each channel and add it to the list
    for channel in range(low_channel, high_channel + 1):
        descriptor = DaqOutChanDescriptor(channel, DaqOutChanType.ANALOG, voltage_range)
        channel_descriptors.append(descriptor)
        amplitudes.append(1.0)  # Volts peak

def configure_digital_channels(daq_device, low_port_index, high_port_index, channel_descriptors, amplitudes):
    """
    Add digital output ports to the channel_descriptors list.

    Raises:
        Exception if a port index is not in range
    """
    dio_device = daq_device.get_dio_device()
    dio_info = dio_device.get_info()
    port_types = dio_info.get_port_types()

    # Validate the low_port_index and high_port_index values
    number_of_ports = len(port_types)
    valid_ports_string = 'valid digital port index values are 0 - {0:d}'.format(number_of_ports - 1)
    if low_port_index < 0 or low_port_index >= number_of_ports:
        error_message = ' '.join(['Error: Invalid digital_low_port_index selection,', valid_ports_string])
        raise Exception(error_message)
    if high_port_index < 0 or high_port_index >= number_of_ports:
        error_message = ' '.join(['Error: Invalid digital_high_port_index selection,', valid_ports_string])
        raise Exception(error_message)

    # Create a channel descriptor for each port and add it to the list
    # Also calculate the amplitude to be used for the digital port waveform
    for port_index in range(low_port_index, high_port_index + 1):
        port = port_types[port_index]

        dio_device.d_config_port(port, DigitalDirection.OUTPUT)
        descriptor = DaqOutChanDescriptor(port, DaqOutChanType.DIGITAL)
        channel_descriptors.append(descriptor)

        port_info = dio_info.get_port_info(port)
        amplitudes.append((pow(2, port_info.number_of_bits) - 1) / 2)


def create_output_data(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, data_buffer, wave):
    shape = wave[0].strip()
    frequency = wave[1] * 2 * np.pi
    shift = wave[2]
    amplitude = wave[3]
    offset = wave[4]
    cycles_per_buffer = int(samples_per_channel / samples_per_cycle)
    i = 0
    for sample in range(samples_per_cycle):
        for chan in channel_descriptors:
            val = 0
            if shape == 'Pulse':
                val = sin(2 * pi * sample / samples_per_cycle*10)*(1-sample/samples_per_cycle)
            elif shape == 'Square':
                val = amplitude * square(frequency * sample / samples_per_cycle + shift)
            elif shape == 'Sawtooth':
                val = amplitude * sawtooth(frequency * sample / samples_per_cycle + shift)
            elif shape == 'Sine':
                val = amplitude * sin(frequency * sample / samples_per_cycle + shift)
            if chan.type == DaqOutChanType.ANALOG:
                data_buffer[i] = val + offset
            else:
                data_buffer[i] = round(val + offset)
            i += 1
            if i >= len(data_buffer):
                return

def create_output_data_xsl(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, data_buffer, file_name):
    book = load_workbook(file_name)
    sheet = book.active
    check = True
    for letter in list(string.ascii_uppercase):
        i = 1
        while check:
            cell = letter + str(i)
            val = sheet[cell].value
            if val is None:
                check = False
                print('last cell placed is ',i-1)
            else:
                if val > 10 or val < -10:
                    val = 0
                data_buffer[i - 1] = float(val)
                print(data_buffer[i-1])
                i += 1
            if i - 1 >= len(data_buffer):
                print('Data input too long! Trimming input data to cell ',i)
                return

def create_output_data_multi(channel_descriptors, samples_per_channel, samples_per_cycle, amplitudes, data_buffer, wave1, wave2):
    shape1 = wave1[0].strip()
    data1 = []
    data2 = []
    if shape1 == "Custom":
        file_name = wave1[1]
        book = load_workbook(file_name)
        sheet = book.active
        check = True
        i = 1
        j = 0
        while check:
            cell = 'A' + str(i)
            val = sheet[cell].value
            if val is None:
                i = 1
            else:
                if val > 10 or val < -10:
                    val = 0
                data1.append(float(val))
                i += 1
                j += 1
            if j >= len(data_buffer) / 2:
                print('Data input too long! Trimming input data to cell ',i)
                check = False
    else:
        frequency1 = wave1[1] * 2 * np.pi
        shift1 = wave1[2]
        amplitude1 = wave1[3]
        offset1 = wave1[4]
        i = 0
        for sample in range(samples_per_channel):
            for chan in channel_descriptors:
                val = 0
                if shape1 == 'Pulse':
                    val = sin(2 * pi * sample / samples_per_cycle*10)*(1-sample/samples_per_cycle)
                elif shape1 == 'Square':
                    val = amplitude1 * square(frequency1 * sample / samples_per_cycle + shift1)
                elif shape1 == 'Sawtooth':
                    val = amplitude1 * sawtooth(frequency1 * sample / samples_per_cycle + shift1)
                elif shape1 == 'Sine':
                    val = amplitude1 * sin(frequency1 * sample / samples_per_cycle + shift1)
                if chan.type == DaqOutChanType.ANALOG:
                    data1.append(val + offset1)
                else:
                    data1.append(round(val + offset1))
                i += 1
                if i >= len(data_buffer) / 2:
                    break
    shape2 = wave2[0].strip()
    if shape2 == "Custom":
        file_name = wave2[1]
        book = load_workbook(file_name)
        sheet = book.active
        check = True
        i = 1
        j = 0
        while check:
            cell = 'B' + str(i)
            val = sheet[cell].value
            if val is None:
                i = 1
            else:
                if val > 10 or val < -10:
                    val = 0
                data2.append(float(val))
                i += 1
                j += 1
            if j >= len(data_buffer) / 2:
                print('Data input too long! Trimming input data to cell ',i)
                check = False
    else:
        frequency2 = wave2[1] * 2 * np.pi
        shift2 = wave2[2]
        amplitude2 = wave2[3]
        offset2 = wave2[4]
        i = 0
        for sample in range(samples_per_channel):
            for chan in channel_descriptors:
                val = 0
                if shape2 == 'Pulse':
                    val = sin(2 * pi * sample / samples_per_cycle*10)*(1-sample/samples_per_cycle)
                elif shape2 == 'Square':
                    val = amplitude2 * square(frequency2 * sample / samples_per_cycle + shift2)
                elif shape2 == 'Sawtooth':
                    val = amplitude2 * sawtooth(frequency2 * sample / samples_per_cycle + shift2)
                elif shape2 == 'Sine':
                    val = amplitude2 * sin(frequency2 * sample / samples_per_cycle + shift2)
                if chan.type == DaqOutChanType.ANALOG:
                    data2.append(val + offset2)
                else:
                    data2.append(round(val + offset2))
                i += 1
                if i >= len(data_buffer) / 2:
                    break
    for i in range(len(data1)):
        data_buffer.append(data1[i])
        data_buffer.append(data2[i])
    return
    
def MCUSB_acquire(high_channel,low_channel):
    DAQ.high_channel = high_channel
    DAQ.low_channel = low_channel
    ai_device = DAQ.ai_device
    samples_per_channel = DAQ.samples_per_channel
    channel_count = high_channel-low_channel+1
    input_mode = DAQ.input_mode
    rate = DAQ.rate
    flags = DAQ.flags
    scan_options = DAQ.scan_options
    ranges = DAQ.ranges
    range_index = DAQ.range_index
    DAQ.buffer_store = create_float_buffer(channel_count,samples_per_channel)

    DAQ.f_out = h5py.File(DAQ.OUTPUT_FILENAME, 'w', libver='latest')
    arr = np.array([np.zeros(channel_count)], dtype='f2')
    DAQ.dset = DAQ.f_out.create_dataset("events", chunks=(DAQ.buffer_store,channel_count), maxshape=(None,None), data=arr, compression="gzip", compression_opts=9)
    DAQ.f_out.swmr_mode = True

    try:
        print('MCUSB_acquire--> commence data aquisition...')
        # Start the acquisition.
        DAQ.rate = ai_device.a_in_scan(low_channel, high_channel, input_mode, ranges[range_index], samples_per_channel, rate, scan_options, flags, DAQ.data)
    except KeyboardInterrupt:
        pass
    except (ValueError, NameError, SyntaxError):
        pass
    except Exception as e:
        print('\n', e)
    
    return


def MCUSB_stop():
    ai_device = DAQ.ai_device
    daq_device = DAQ.daq_device
    ai_info = DAQ.ai_device.get_info()
    if daq_device:
        if ai_device and ai_info and ai_info.has_pacer():
            #os.system('clear')
            print('MCUSB_stop--> ai_device.scan_stop()...')
            ai_device.scan_stop()
            DAQ.f_out.close() # close the hdf5 file
        # daq_device.disable_event(DAQ.event_types)
        # if daq_device.is_connected():
            # daq_device.disconnect()
        # daq_device.release()
    return

def event_callback_function(event_callback_args):
    # type: (EventCallbackArgs) -> None
    """
    The callback function called in response to an event condition.

    Args:
        event_callback_args: Named tuple :class:`EventCallbackArgs` used to pass
            parameters to the user defined event callback function :class`DaqEventCallback`.
            The named tuple contains the following members
            event_type - the condition that triggered the event
            event_data - additional data that specifies an event condition
            user_data - user specified data that will be passed to the callback function
    """

    event_type = event_callback_args.event_type
    event_data = event_callback_args.event_data
    user_data = event_callback_args.user_data

    if event_type == DaqEventType.ON_DATA_AVAILABLE:
        #reset_cursor()
        #os.system('clear')
        #print('Hit ENTER to stop data acquisition')
        print('\n')
        #print('eventType: ', DaqEventType(event_type).name)

        scan_event_parameters = user_data
        total_events = event_data
        chan_count = scan_event_parameters.high_chan - scan_event_parameters.low_chan + 1
        buffer_len = len(scan_event_parameters.buffer)

        # Keep track of index, even with a wrap around buffer
        startIndex = ((total_events - user_data.buffer_store) * chan_count) % buffer_len
        endIndex = (total_events * chan_count) % buffer_len

        # Store to HDF5 File -----------------------------------
        if (endIndex < startIndex):
            data = np.append(scan_event_parameters.buffer[startIndex:], scan_event_parameters.buffer[:endIndex])
        else:  
            data = scan_event_parameters.buffer[startIndex:endIndex]

        data = np.reshape(data,(-1,chan_count))
        dfIndex = np.arange(total_events - user_data.buffer_store, total_events, dtype=int)
        #df= pd.DataFrame(data, columns=DAQ.columnNames[:chan_count], index=dfIndex)
        #df.to_hdf(DAQ.OUTPUT_FILENAME, key=DAQ.hdfKey, format='t', data_columns=True, mode='a', append=True, complib=DAQ.comp_lib, complevel=DAQ.comp_level)
    
        new_shape = (total_events, chan_count)
        DAQ.dset.resize( new_shape )
        findex = total_events -user_data.buffer_store
        DAQ.dset[findex:,:] = data  #dset[...] = data # writing data to the output file
        DAQ.dset.flush()

        # Print outputs
        print('Event counts (total): ', total_events)
        print('Scan rate = ', '{:.2f}'.format(DAQ.rate), 'Hz')
        print('buffer_length = ', buffer_len)
        print('currentBufferIndex = ', startIndex)
        for i in range(chan_count):
            print('chan',
                  i + scan_event_parameters.low_chan,
                  '{:.6f}'.format(scan_event_parameters.buffer[endIndex - chan_count + i]))

    if event_type == DaqEventType.ON_INPUT_SCAN_ERROR:
        exception = ULException(event_data)
        print(exception)

    if event_type == DaqEventType.ON_END_OF_INPUT_SCAN:
        print('\nThe scan is complete, hit ENTER to continue\n')

    return
    
def display_scan_options(bit_mask):

    options = []
    if bit_mask == ScanOption.DEFAULTIO:
        options.append(ScanOption.DEFAULTIO.name)
    for so in ScanOption:
        if so & bit_mask:
            options.append(so.name)
    return ', '.join(options)

def reset_cursor():
    stdout.write('\033[1;1H')
    
def get_analog_channels():
    ai_device = DAQ.ai_device.get_info()
    input_chans = ai_device.get_num_chans()
    ao_device = DAQ.ao_device.get_info()
    output_chans = ao_device.get_num_chans()
    input_list = []
    output_list = []
    for i in range(input_chans):
        string = 'Input ' + str(i)
        input_list.append(string)
    for i in range(output_chans):
        string = 'Output ' + str(i)
        output_list.append(string)
    return [input_list,output_list]
    

context = zmq.Context()
socket = context.socket(zmq.REP)
#socket = context.socket(zmq.DEALER)
socket.bind("tcp://*:5555")

MCUSB_initiate() # initiate the connection to the MCUSB box

while True:
    #  Wait for next request from client
    message = socket.recv().decode()
    #os.system('clear')
    if message=='Alive?':
        socket.send(b"Server is ready...")
    else:
        print("0mq: Received request: %s" % message)
    if message == 'Acquire Channels':
        socket.send(str(get_analog_channels()).encode())
    elif message == 'Stop':
        socket.send(b"Reply: MCUSB stops acquisition")
        MCUSB_stop()
    elif message == 'Stop Wave':
        socket.send(b"Reply: Stopping previous wave")
        MCUSB_stop_ao()
    waveform_data = message.strip('][').replace('\'','').replace(' ','').split(',')
    if len(waveform_data) > 1:
        if waveform_data[0] == 'True':
            waveform_data.pop(0)
            if waveform_data[0] == 'Custom':
                if waveform_data[2] == 'Custom':
                    # Both are custom
                    socket.send(b"Reply: Forming two custom waves.")
                    MCUSB_multi_ao("Custom","Custom",waveform_data)
                else:
                    # Only the first is custom
                    socket.send(b"Reply: Forming a custom and %s wave" % waveform_data[2].encode())
                    MCUSB_multi_ao("Custom",waveform_data[2],waveform_data)
            elif waveform_data[5] == 'Custom':
                # Only the second is custom
                socket.send(b"Reply: Forming a %s and a custom wave" % waveform_data[0].encode())
                MCUSB_multi_ao(waveform_data[0],"Custom",waveform_data)
            else:
                # Both are precoded
                socket.send(b"Reply: Forming a %s and a %s wave" % (waveform_data[0].encode(), waveform_data[5].encode()))
                MCUSB_multi_ao(waveform_data[0],waveform_data[5],waveform_data)
        elif waveform_data[0] == 'False':
            socket.send(b"Reply: Forming %s wave!" % waveform_data[1].encode())
            waveform_data.pop(0)
            MCUSB_ao(waveform_data)
        elif waveform_data[0] == 'Digital':
            waveform_data.pop(0)
            if waveform_data[0] == 'True':
                waveform_data.pop(0)
                if waveform_data[0] == 'Custom':
                    if waveform_data[2] == 'Custom':
                        # Both are custom
                        socket.send(b"Reply: Forming two custom waves.")
                        MCUSB_multi_do("Custom","Custom",waveform_data)
                    else:
                        # Only the first is custom
                        socket.send(b"Reply: Forming a custom and %s wave" % waveform_data[2].encode())
                        MCUSB_multi_do("Custom",waveform_data[2],waveform_data)
                elif waveform_data[5] == 'Custom':
                    # Only the second is custom
                    socket.send(b"Reply: Forming a %s and a custom wave" % waveform_data[0].encode())
                    MCUSB_multi_do(waveform_data[0],"Custom",waveform_data)
                else:
                    # Both are precoded
                    socket.send(b"Reply: Forming a %s and a %s wave" % (waveform_data[0].encode(), waveform_data[5].encode()))
                    MCUSB_multi_do(waveform_data[0],waveform_data[5],waveform_data)
            elif waveform_data[0] == 'False':
                socket.send(b"Reply: Forming %s wave!" % waveform_data[1].encode())
                waveform_data.pop(0)
                MCUSB_do(waveform_data)
        elif type(waveform_data[0]) == str and waveform_data[0][0:7] == 'Acquire':
            socket.send(b"Reply: MCUSB starts acquisition")
            file_no = waveform_data[0].strip('Acquire')
            DAQ.OUTPUT_FILENAME= 'test'+file_no+'.hdf'
            MCUSB_acquire(int(waveform_data[1])+1,int(waveform_data[2]))
    
